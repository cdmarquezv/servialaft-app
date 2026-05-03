import streamlit as st
import pandas as pd
import unicodedata
import io
import json as _json
import os as _os
import urllib.parse as _uparse
import xml.etree.ElementTree as _ET
import urllib.request as _ureq
from datetime import datetime, date

try:
    from rapidfuzz import fuzz
except ImportError:
    st.error("Instala rapidfuzz: pip install rapidfuzz"); st.stop()

try:
    from generador_pdf import generar_pdf_individual, generar_pdf_manual, generar_pdf_masivo
    PDF_DISPONIBLE = True
except ImportError:
    PDF_DISPONIBLE = False

from database import (
    init_db, verificar_usuario,
    listar_empresas, crear_empresa, toggle_empresa,
    listar_usuarios, crear_usuario, toggle_usuario, reset_password,
    registrar_consulta, listar_consultas,
)
init_db()

st.set_page_config(page_title="SERVIALAFT SAS", page_icon="🛡️",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Domine:wght@400;700&display=swap');

/* ── Sidebar ─────────────────────────────────────────────────── */
[data-testid="stSidebar"] { background:#0F3D3A; }
[data-testid="stSidebar"] * { color:#e2eeec !important; }
[data-testid="stSidebar"] .stRadio > label { display:none; }
[data-testid="stSidebar"] .stRadio label {
    display:flex !important; align-items:center; padding:10px 14px;
    border-radius:8px; margin:2px 0; cursor:pointer; font-size:15px; }
[data-testid="stSidebar"] .stRadio label:hover { background:rgba(65,189,178,0.18); }
[data-testid="stSidebar"] .stRadio label:has(input:checked) {
    background:rgba(65,189,178,0.28) !important;
    border-left:3px solid #41BDB2; }

/* ── Botones primarios ───────────────────────────────────────── */
.stButton > button[kind="primary"],
button[data-testid="baseButton-primary"] {
    background:#009881 !important;
    border:none !important;
    color:#fff !important;
    border-radius:8px !important; }
.stButton > button[kind="primary"]:hover,
button[data-testid="baseButton-primary"]:hover {
    background:#0F3D3A !important; }

/* ── Login ───────────────────────────────────────────────────── */
.login-wrap { max-width:420px; margin:60px auto 0; padding:40px 36px 32px;
    background:#fff; border-radius:16px; box-shadow:0 4px 32px rgba(15,61,58,0.12); }
.login-title { font-size:26px; font-weight:700; font-family:'Domine',serif;
    color:#0F3D3A; text-align:center; margin-bottom:4px; }
.login-sub   { font-size:13px; color:#51535B; text-align:center; margin-bottom:24px; }

/* ── Tarjetas de fuentes externas ────────────────────────────── */
.ext-link { display:block; padding:12px 16px; border-radius:10px;
    border:1px solid #b2d8d5; background:#f0faf9;
    text-decoration:none !important; margin-bottom:8px; transition:box-shadow .2s; }
.ext-link:hover { box-shadow:0 4px 12px rgba(15,61,58,0.12); }
.ext-title { font-size:14px; font-weight:700; color:#0F3D3A; }
.ext-url   { font-size:11px; color:#009881; margin-top:3px; font-family:monospace; }

/* ── Métricas ────────────────────────────────────────────────── */
.metric-box { background:#f0faf9; border:1px solid #b2d8d5;
    border-radius:12px; padding:20px 16px; text-align:center; }
.metric-num   { font-size:38px; font-weight:800; font-family:'Domine',serif;
    color:#0F3D3A; line-height:1; }
.metric-label { font-size:13px; color:#51535B; margin-top:6px; }

/* ── Headings principales ────────────────────────────────────── */
h1, h2, h3, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
    font-family:'Domine',serif !important; color:#0F3D3A !important; }

/* ── Footer ──────────────────────────────────────────────────── */
.footer-brand {
    position:fixed; bottom:0; left:0; right:0;
    background:#0F3D3A; color:#a8ccc9;
    text-align:center; padding:8px 16px;
    font-size:12px; z-index:999;
    border-top:2px solid #41BDB2;
}
.footer-brand a { color:#41BDB2; text-decoration:none; }
.footer-brand a:hover { text-decoration:underline; }
footer { visibility:hidden; } #MainMenu { visibility:hidden; }
</style>
""", unsafe_allow_html=True)

# Footer de marca — siempre visible
st.markdown("""
<div class="footer-brand">
    🛡️ <b>CruzaListas</b> &nbsp;·&nbsp;
    Desarrollado y operado por <b>SERVIALAFT SAS</b> &nbsp;·&nbsp;
    Todos los derechos reservados © 2025 &nbsp;·&nbsp;
    <a href="mailto:contacto@servialaft.com">contacto@servialaft.com</a>
</div>
""", unsafe_allow_html=True)

for k, v in {
    "logged_in": False,
    "user": None,
    "user_info": {},
    "menu": "📊 Estadísticas de uso",
    "ultima_consulta": {"tipo_id": "CC", "nro_id": "", "nombre": ""},
    "login_intentos": 0,
    "login_bloqueado_hasta": None,
}.items():
    if k not in st.session_state: st.session_state[k] = v

def cargar_listas():
    if _os.path.exists("listas_vinculantes.json"):
        with open("listas_vinculantes.json","r",encoding="utf-8") as f:
            data = _json.load(f)
        rows = []
        for r in data.get("registros",[]):
            nom=r.get("nombre",""); lst=r.get("lista","OFAC SDN")
            prog=", ".join(r.get("programas",[])) or lst
            det=r.get("detalle") or prog
            rows.append({"tipo_id":"N/A","nro_id":"N/A","nombre":nom,"origen":lst,"detalle":det})
            for doc in r.get("documentos",[]):
                num=doc.get("numero","").strip()
                if num and num!="N/A":
                    rows.append({"tipo_id":doc.get("tipo","DOC"),"nro_id":num,"nombre":nom,"origen":lst,"detalle":det})
            for aka in r.get("aka",[]):
                if aka: rows.append({"tipo_id":"N/A","nro_id":"N/A","nombre":aka,"origen":lst,"detalle":f"{det} (AKA)"})
        return pd.DataFrame(rows), True, data.get("meta",{})
    # ── DATOS DEMO — Colombianos reales en listas vinculantes ────────────────
    demo = pd.DataFrame([
        # OFAC SDN — Narcotráfico Colombia
        {"tipo_id":"CC", "nro_id":"19400625","nombre":"PABLO ESCOBAR GAVIRIA","origen":"OFAC SDN","detalle":"NARCOTICS — Cartel de Medellín"},
        {"tipo_id":"CC", "nro_id":"16354146","nombre":"GILBERTO RODRIGUEZ OREJUELA","origen":"OFAC SDN","detalle":"NARCOTICS — Cartel de Cali"},
        {"tipo_id":"CC", "nro_id":"16354147","nombre":"MIGUEL ANGEL RODRIGUEZ OREJUELA","origen":"OFAC SDN","detalle":"NARCOTICS — Cartel de Cali"},
        {"tipo_id":"CC", "nro_id":"71612512","nombre":"DAIRO ANTONIO USUGA DAVID","origen":"OFAC SDN","detalle":"NARCOTICS — Clan del Golfo (Otoniel)"},
        {"tipo_id":"CC", "nro_id":"98530916","nombre":"JOSE GONZALO RODRIGUEZ GACHA","origen":"OFAC SDN","detalle":"NARCOTICS — El Mexicano"},
        {"tipo_id":"CC", "nro_id":"71700551","nombre":"DANIEL RENDON HERRERA","origen":"OFAC SDN","detalle":"NARCOTICS — Don Mario"},
        {"tipo_id":"CC", "nro_id":"77040924","nombre":"HENRY DE JESUS LOPEZ LONDONO","origen":"OFAC SDN","detalle":"NARCOTICS — Mi Sangre"},
        {"tipo_id":"CC", "nro_id":"79945621","nombre":"GUSTAVO FRANCISCO PETRO URREGO","origen":"OFAC SDN","detalle":"NARCOTICS — Lista Clinton · Incluido Oct 2025 por Trump · Vínculos narcoterrorismo"},
        {"tipo_id":"NIT","nro_id":"800154059","nombre":"GRUPO EMPRESARIAL DAABON","origen":"OFAC SDN","detalle":"SDGT — Vínculos narcoactividad"},
        {"tipo_id":"NIT","nro_id":"900234567","nombre":"ODEBRECHT COLOMBIA","origen":"OFAC SDN","detalle":"SDGT — Sobornos y corrupción"},
        # ONU — Sanciones
        {"tipo_id":"CC", "nro_id":"19430101","nombre":"IVAN MARQUEZ","origen":"ONU","detalle":"RES. 2341 — FARC disidencias"},
        {"tipo_id":"CC", "nro_id":"19620303","nombre":"JESUS SANTRICH","origen":"ONU","detalle":"RES. 2341 — FARC narcotrafico"},
        # Terroristas
        {"tipo_id":"CC", "nro_id":"71359678","nombre":"TIMOLEON JIMENEZ","origen":"TERRORISTAS EE.UU.","detalle":"SDGT — FARC Timochenko"},
        {"tipo_id":"CC", "nro_id":"17030128","nombre":"RODRIGO LONDONO ECHEVERRI","origen":"TERRORISTAS EE.UU.","detalle":"SDGT — FARC Timochenko (AKA)"},
    ])
    return demo, False, {}

TODAS, LISTAS_REALES, LISTAS_META = cargar_listas()

def norm(s):
    if not isinstance(s,str): return ""
    return "".join(c for c in unicodedata.normalize("NFD",s) if unicodedata.category(c)!="Mn").upper().strip()

def buscar(tipo_id, nro_id, nombre, umbral):
    res=[]; nombre_n=norm(nombre)
    solo_doc=bool(nro_id) and not nombre
    solo_nombre=bool(nombre) and not nro_id
    for _,row in TODAS.iterrows():
        tiene_doc=(str(row.get("nro_id","N/A")) not in ("N/A","") and str(row.get("tipo_id","N/A"))!="N/A")
        sim=0.0; nivel=""
        if solo_doc:
            if not tiene_doc: continue
            if row["tipo_id"]!=tipo_id: continue
            if str(row["nro_id"])==str(nro_id): nivel="EXACTA"
            elif fuzz.ratio(str(row["nro_id"]),str(nro_id))>=85: nivel="APROXIMADA"
            else: continue
            sim=100.0
        elif solo_nombre:
            sim=fuzz.token_sort_ratio(norm(str(row["nombre"])),nombre_n)/100
            if sim<umbral: continue
            nivel="SOLO NOMBRE"
        else:
            sim=fuzz.token_sort_ratio(norm(str(row["nombre"])),nombre_n)/100
            de=dc=False
            if tiene_doc and row["tipo_id"]==tipo_id:
                de=str(row["nro_id"])==str(nro_id)
                dc=fuzz.ratio(str(row["nro_id"]),str(nro_id))>=85
            if de and sim>=umbral: nivel="EXACTA"
            elif dc and sim>=umbral: nivel="APROXIMADA"
            elif sim>=umbral: nivel="SOLO NOMBRE"
            else: continue
        r=row.to_dict(); r["sim_%"]=round(sim*100,1) if sim<=1 else round(sim,1); r["nivel"]=nivel
        res.append(r)
    seen=set(); out=[]
    for r in res:
        k=(r["nombre"],r["origen"])
        if k not in seen: seen.add(k); out.append(r)
    return pd.DataFrame(out) if out else pd.DataFrame()

def log_q(modulo, tipo_id, nro_id, nombre, resultado, empresa_consultada=""):
    info = st.session_state.get("user_info", {})
    registrar_consulta(
        username=st.session_state.user,
        empresa_id=info.get("empresa_id", 1),
        modulo=modulo,
        tipo_id=tipo_id,
        nro_id=str(nro_id),
        nombre=nombre,
        resultado=resultado,
        empresa_consultada=empresa_consultada,
    )

def a_excel(df):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        df.to_excel(w,index=False,sheet_name="Datos")
        ws=w.sheets["Datos"]
        from openpyxl.styles import Font,PatternFill,Alignment
        for cell in ws[1]:
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor="0F1B2D")
            cell.alignment=Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=max(len(str(col[0].value or "")),10)+4
    return buf.getvalue()

PALABRAS_RIESGO=["lavado de activos","narcotrafico","corrupcion","fraude","capturado","investigado",
    "condenado","sancionado","terrorismo","extorsion","contrabando","enriquecimiento ilicito",
    "peculado","cohecho","imputado","money laundering","fraud","corruption","arrested","sanctioned",
    "drug trafficking","terrorism","formulacion de cargos"]

def _rss_fetch(url,max_n=8):
    try:
        req=_ureq.Request(url,headers={"User-Agent":"Mozilla/5.0"})
        with _ureq.urlopen(req,timeout=10) as resp:
            root=_ET.fromstring(resp.read())
        items=[]
        for item in root.findall(".//item")[:max_n]:
            titulo=item.findtext("title","").strip()
            if not titulo: continue
            items.append({"titulo":titulo,"fuente":item.findtext("source","").strip(),
                "fecha":(item.findtext("pubDate","") or "")[:16],
                "link":item.findtext("link","").strip(),
                "riesgo":any(p in norm(titulo).lower() for p in PALABRAS_RIESGO)})
        return items,None
    except Exception as e: return [],str(e)

def buscar_noticias(nombre, pais="Colombia", max_n=8):
    """Busca noticias adversas usando nombre completo entre comillas para evitar falsos positivos."""
    # Nombre completo entre comillas = exige coincidencia exacta de frase
    nombre_q = f'"{nombre}"'
    t = "lavado OR narcotrafico OR corrupcion OR fraude OR capturado OR investigado OR condenado OR sancionado OR terrorismo OR imputado OR extorsion OR peculado"
    q = f'{nombre_q} ({t})'
    if pais: q += f" {pais}"
    return _rss_fetch(
        f"https://news.google.com/rss/search?q={_uparse.quote(q)}&hl=es-419&gl=CO&ceid=CO:es-419",
        max_n
    )

def buscar_noticias_fiscalia(nombre, max_n=6):
    """Busca directamente en la Fiscalía con nombre completo exacto."""
    resultados = []
    # Método 1: RSS WordPress de fiscalia.gov.co con nombre entre comillas
    try:
        url_rss = f"https://www.fiscalia.gov.co/colombia/?s={_uparse.quote(f'{nombre}')}&feed=rss2"
        req = _ureq.Request(url_rss, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        })
        with _ureq.urlopen(req, timeout=12) as resp:
            root = _ET.fromstring(resp.read())
        for item in root.findall(".//item")[:max_n]:
            titulo = item.findtext("title","").strip()
            link   = item.findtext("link","").strip()
            fecha  = (item.findtext("pubDate","") or "")[:16]
            desc   = item.findtext("description","").strip()[:200]
            if titulo:
                # Verificar que el título contenga al menos parte del nombre buscado
                partes_nombre = [p for p in nombre.upper().split() if len(p) > 3]
                coincide = any(p in titulo.upper() for p in partes_nombre[:2])
                if coincide:
                    resultados.append({
                        "titulo": titulo, "fuente": "Fiscalía General de la Nación",
                        "fecha": fecha, "link": link, "desc": desc,
                        "origen": "Fiscalía", "riesgo": True,
                    })
    except Exception:
        pass
    # Método 2: Google News con nombre exacto + fiscalía (fallback)
    if not resultados:
        try:
            q = f'"{nombre}" site:fiscalia.gov.co OR "{nombre}" fiscalia colombia imputado OR capturado OR condenado'
            url_gn = f"https://news.google.com/rss/search?q={_uparse.quote(q)}&hl=es-419&gl=CO&ceid=CO:es-419"
            req2 = _ureq.Request(url_gn, headers={"User-Agent": "Mozilla/5.0"})
            with _ureq.urlopen(req2, timeout=10) as resp:
                root2 = _ET.fromstring(resp.read())
            for item in root2.findall(".//item")[:max_n]:
                titulo = item.findtext("title","").strip()
                link   = item.findtext("link","").strip()
                fecha  = (item.findtext("pubDate","") or "")[:16]
                if titulo and nombre.split()[0].upper() in titulo.upper():
                    resultados.append({
                        "titulo": titulo, "fuente": "Fiscalía (vía Google)",
                        "fecha": fecha, "link": link, "desc": "",
                        "origen": "Fiscalía", "riesgo": True,
                    })
        except Exception:
            pass
    return resultados

def mostrar_noticias(nom_label, tipo_id, id_label):
    nombre_enc = _uparse.quote(nom_label)
    url_fiscalia = f"https://www.fiscalia.gov.co/colombia/?s={nombre_enc}"

    st.markdown("### 📰 Noticias adversas y fuentes judiciales")

    # Link directo Fiscalía con nombre precargado
    st.markdown(f"""
    <a class="ext-link" href="{url_fiscalia}" target="_blank">
      <div class="ext-title">🏛️ Buscar en Fiscalía General de la Nación</div>
      <div style="font-size:12px;color:#374151;margin-top:2px">
        Búsqueda directa para: <b>{nom_label}</b>
      </div>
      <div class="ext-url">fiscalia.gov.co/colombia/?s={nom_label}</div>
    </a>
    """, unsafe_allow_html=True)

    col_g, col_f = st.columns(2)

    with col_g:
        st.markdown("**🌐 Google News — Noticias adversas**")
        with st.spinner("Buscando noticias..."):
            noticias, err = buscar_noticias(nom_label, "Colombia", 8)
        if err:
            st.caption(f"No disponible: {err}")
        elif not noticias:
            st.success("✅ Sin noticias adversas encontradas")
        else:
            adv = [n for n in noticias if n["riesgo"]]
            neu = [n for n in noticias if not n["riesgo"]]
            if adv:
                st.error(f"🚨 {len(adv)} noticia(s) con indicadores de riesgo")
                log_q("NOTICIAS", tipo_id, id_label, nom_label, f"{len(adv)} NOTICIAS ADVERSAS")
            else:
                st.info(f"📰 {len(noticias)} noticias — sin riesgo directo")
            for n in (adv + neu):
                e = "🔴" if n["riesgo"] else "🟡"
                with st.expander(f"{e} {n['titulo'][:75]}",
                                 expanded=n["riesgo"]):
                    c1, c2 = st.columns([2, 1])
                    c1.caption(f"📰 {n['fuente']}")
                    c2.caption(f"📅 {n['fecha']}")
                    st.markdown(f"[🔗 Ver noticia completa]({n['link']})")

    with col_f:
        st.markdown("**🏛️ Fiscalía — Resultados directos**")
        with st.spinner("Consultando Fiscalía..."):
            nf = buscar_noticias_fiscalia(nom_label)
        if not nf:
            st.success("✅ Sin resultados en Fiscalía General")
            st.markdown(f"[🔍 Verificar manualmente]({url_fiscalia})")
        else:
            st.error(f"🚨 {len(nf)} resultado(s) en Fiscalía General")
            for n in nf:
                with st.expander(f"🔴 {n['titulo'][:75]}", expanded=True):
                    c1, c2 = st.columns([2, 1])
                    c1.caption(f"🏛️ {n['fuente']}")
                    c2.caption(f"📅 {n['fecha']}")
                    if n.get("desc"):
                        st.caption(n["desc"] + "…")
                    st.markdown(f"[🔗 Ver en Fiscalía.gov.co]({n['link']})")
            st.markdown(f"[🔍 Ver todos los resultados en Fiscalía]({url_fiscalia})")

def pantalla_login():
    from datetime import timedelta
    _,col,_=st.columns([1,1.3,1])
    with col:
        st.markdown("""<div class="login-wrap">
          <div class="login-title">🛡️ CruzaListas</div>
          <div class="login-sub">Sistema de Consulta de Listas Vinculantes<br>
          OFAC · ONU · Terroristas · PEPs · Noticias Adversas<br><br>
          <span style="font-size:11px;color:#9ca3af;">
          Desarrollado y operado por <b>SERVIALAFT SAS</b>
          </span></div>
        </div>""",unsafe_allow_html=True)

        bloqueado_hasta = st.session_state.login_bloqueado_hasta
        if bloqueado_hasta and datetime.now() < bloqueado_hasta:
            segundos = int((bloqueado_hasta - datetime.now()).total_seconds())
            st.error(f"Demasiados intentos fallidos. Espera {segundos} segundo(s).")
            return

        usuario  = st.text_input("Usuario", placeholder="usuario")
        password = st.text_input("Contraseña", type="password", placeholder="contraseña")
        if st.button("Iniciar sesión →", type="primary", use_container_width=True):
            info = verificar_usuario(usuario, password)
            if info:
                st.session_state.logged_in  = True
                st.session_state.user       = usuario
                st.session_state.user_info  = info
                st.session_state.login_intentos = 0
                st.session_state.login_bloqueado_hasta = None
                st.rerun()
            else:
                st.session_state.login_intentos += 1
                intentos = st.session_state.login_intentos
                if intentos >= 5:
                    st.session_state.login_bloqueado_hasta = datetime.now() + timedelta(minutes=5)
                    st.error("Demasiados intentos. Bloqueado por 5 minutos.")
                else:
                    st.error(f"Usuario o contraseña incorrectos. {5-intentos} intento(s) restante(s).")

MENU_ANALISTA = ["📊 Estadísticas de uso",
                 "🔍 Búsqueda Unificada", "🔗 Otras Fuentes",
                 "📋 Registros consultados",
                 "🔑 Mi perfil",
                 "🚪 Cerrar sesión"]
MENU_SUPER    = ["📊 Estadísticas de uso",
                 "🔍 Búsqueda Unificada", "🔗 Otras Fuentes",
                 "📋 Registros consultados",
                 "🏢 Empresas", "👥 Usuarios",
                 "🔑 Mi perfil",
                 "🚪 Cerrar sesión"]

def sidebar():
    info = st.session_state.get("user_info", {})
    es_super = info.get("rol") == "superadmin"
    menu_items = MENU_SUPER if es_super else MENU_ANALISTA

    with st.sidebar:
        st.markdown("### 🛡️ CruzaListas")
        st.caption("por SERVIALAFT SAS")
        st.markdown("---")
        st.markdown(f"**{info.get('nombre', st.session_state.user)}**")
        st.caption(f"{'Superadministrador' if es_super else 'Analista'} · {info.get('empresa_nombre','')}")
        st.markdown("---")
        idx = 0
        for i, m in enumerate(menu_items):
            if st.session_state.menu in m or m in st.session_state.menu:
                idx = i; break
        sel = st.radio("Menú", menu_items, index=idx, label_visibility="collapsed")
        if "Cerrar" in sel:
            st.session_state.logged_in = False
            st.session_state.user      = None
            st.session_state.user_info = {}
            st.session_state.menu      = "UNIFICADA"
            st.rerun()
        else:
            st.session_state.menu = sel

def mod_unificada():
    st.markdown("## 🔍 Búsqueda Unificada")
    st.caption("OFAC SDN · ONU · Terroristas EE.UU. · Terroristas UE · PEPs Colombia · Noticias Adversas · Fiscalía")
    if LISTAS_REALES:
        total=LISTAS_META.get("total",len(TODAS)); fecha=LISTAS_META.get("fecha_actualizacion","—")
        st.success(f"✅ Listas reales — **{total:,} registros** | Actualización: {fecha}")
    else:
        st.warning("⚠️ Datos de demo — corre `descargar_listas.py` para activar datos reales (incluye PEPs desde datos.gov.co).")

    tab_ind,tab_mas=st.tabs(["🔎 Consulta individual","📂 Carga masiva (Excel)"])

    with tab_ind:
        c1,c2=st.columns([1,2])
        with c1:
            uc = st.session_state.ultima_consulta
            tipo_id=st.selectbox("Tipo ID",["CC","NIT","CE","PAS"],
                index=["CC","NIT","CE","PAS"].index(uc["tipo_id"]) if uc["tipo_id"] in ["CC","NIT","CE","PAS"] else 0)
            nro_id=st.text_input("Número de identificación",value=uc["nro_id"],placeholder="Opcional")
            nombre=st.text_input("Nombre completo",value=uc["nombre"],placeholder="Opcional")
            st.caption("💡 Busca por nombre, documento o ambos.")
            umbral=st.slider("SimiliScore™ (%)",50,100,85)
            consultar=st.button("🔍 Consultar",type="primary",use_container_width=True)
            st.markdown("---")
            st.caption("💡 Para Policía, Procuraduría, Rama Judicial y otras fuentes ve a **🔗 Otras Fuentes**.")
        with c2:
            if consultar:
                td=bool(nro_id.strip()); tn=bool(nombre.strip())
                if not td and not tn:
                    st.warning("Ingresa al menos un número de identificación o un nombre.")
                else:
                    il=nro_id.strip() if td else "N/A"
                    nl=nombre.strip() if tn else "(búsqueda por documento)"
                    # Guardar en session_state para precargar Policía/Procuraduría
                    st.session_state.ultima_consulta = {
                        "tipo_id": tipo_id, "nro_id": nro_id.strip(), "nombre": nombre.strip()
                    }
                    with st.spinner("Consultando listas vinculantes..."):
                        df=buscar(tipo_id,nro_id.strip() if td else "",nombre.strip() if tn else "",umbral/100)
                    st.markdown("### 📋 Listas vinculantes")
                    if df.empty:
                        st.success("✅ Sin coincidencias en ninguna lista vinculante.")
                        log_q("UNIFICADA",tipo_id,il,nl,"NO ENCONTRADO")
                    else:
                        lh=df["origen"].unique().tolist()
                        st.error(f"🚨 Encontrado en {len(lh)} lista(s) — {len(df)} coincidencia(s)")
                        log_q("UNIFICADA",tipo_id,il,nl,df["nivel"].iloc[0])
                        cr=st.columns(min(len(lh),4))
                        for i,lst in enumerate(lh): cr[i%4].metric(lst,f"{len(df[df['origen']==lst])} hit(s)")
                        st.markdown("---")
                        for _,row in df.iterrows():
                            e="🔴" if row["nivel"]=="EXACTA" else "🟡"
                            with st.expander(f"{e} {row['nombre']}  —  {row['origen']}",expanded=True):
                                cc1,cc2,cc3=st.columns(3)
                                cc1.metric("Lista",row["origen"]); cc2.metric("SimiliScore™",f"{row['sim_%']}%"); cc3.metric("Nivel",row["nivel"])
                                if row.get("detalle"): st.write(f"**Detalle:** {row['detalle']}")
                    if tn:
                        st.markdown("---")
                        # Guardar noticias en session para pasarlas al PDF
                        with st.spinner("Buscando noticias adversas..."):
                            _not_g, _  = buscar_noticias(nl, "Colombia", 8)
                            _not_f     = buscar_noticias_fiscalia(nl)
                        st.session_state["_last_noticias_g"] = _not_g
                        st.session_state["_last_noticias_f"] = _not_f
                        mostrar_noticias(nl, tipo_id, il)
                    else:
                        st.session_state["_last_noticias_g"] = []
                        st.session_state["_last_noticias_f"] = []
                    st.markdown("---")
                    cx, cy = st.columns(2)
                    de = df if not df.empty else pd.DataFrame({
                        "tipo_id":[tipo_id],"nro_id":[il],"nombre":[nl],"resultado":["SIN COINCIDENCIA"]})
                    cx.download_button("📥 Excel", data=a_excel(de),
                        file_name=f"consulta_{il}_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                    if PDF_DISPONIBLE:
                        pb = generar_pdf_individual(
                            tipo_id, il, nl,
                            df if not df.empty else None,
                            usuario=st.session_state.user,
                            noticias_google=st.session_state.get("_last_noticias_g",[]),
                            noticias_fiscalia=st.session_state.get("_last_noticias_f",[]),
                            watermark=not LISTAS_REALES,
                        )
                        cy.download_button("📄 Certificado PDF", data=pb,
                            file_name=f"certificado_{il}_{date.today()}.pdf",
                            mime="application/pdf", use_container_width=True)
            else:
                st.info("Completa el formulario y presiona **Consultar**.")
                st.markdown("""
                **Esta búsqueda consulta simultáneamente:**
                - 🇺🇸 OFAC SDN — Nacionales Especialmente Designados
                - 🇺🇸 Terroristas EE.UU. — SDGT, TALIBAN y otros programas
                - 🇺🇳 ONU — Lista Consolidada Consejo de Seguridad
                - 🇪🇺 Terroristas UE — Lista Consolidada de Sanciones
                - 🇨🇴 PEPs Colombia — Personas Expuestas Políticamente
                - 📰 Google News — noticias adversas automático
                - 🏛️ Fiscalía General — noticias automático
                """)

    with tab_mas:
        c1,c2=st.columns([1,2])
        with c1:
            st.markdown("**Formato requerido:**"); st.code("tipo_id | nro_id | nombre")
            pt=pd.DataFrame({"tipo_id":["CC","NIT","CE"],"nro_id":["12345678","900123456","87654321"],
                             "nombre":["JUAN PEREZ GOMEZ","EMPRESA ABC SAS","CARLOS LOPEZ"]})
            st.download_button("📄 Plantilla",data=a_excel(pt),file_name="plantilla.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            archivo=st.file_uploader("Subir .xlsx",type=["xlsx"])
            umbral_m=st.slider("SimiliScore™ masivo (%)",50,100,85)
            _info_m = st.session_state.get("user_info", {})
            _es_super_m = _info_m.get("rol") == "superadmin"
            emp_consultada_m = ""
            if _es_super_m:
                emp_consultada_m = st.text_input("Empresa consultada (cliente)",
                    placeholder="Nombre de la empresa que genera la consulta", key="emp_cons_masivo")
            procesar=st.button("⚙️ Procesar archivo",type="primary",use_container_width=True)
        with c2:
            if procesar and archivo:
                di=pd.read_excel(archivo); di.columns=[c.lower().strip() for c in di.columns]
                if not {"tipo_id","nro_id","nombre"}.issubset(di.columns):
                    st.error("El archivo debe tener columnas: tipo_id, nro_id, nombre")
                else:
                    res=[]; prog=st.progress(0)
                    for i,row in di.iterrows():
                        prog.progress((i+1)/len(di),f"Procesando {i+1}/{len(di)}…")
                        dr=buscar(str(row["tipo_id"]).upper().strip(),str(row["nro_id"]).strip(),str(row["nombre"]).strip(),umbral_m/100)
                        if dr.empty:
                            res.append({"tipo_id":row["tipo_id"],"nro_id":row["nro_id"],"nombre":row["nombre"],
                                        "resultado":"SIN COINCIDENCIA","origen":"—","nivel":"—","sim_%":"—"})
                            log_q("MASIVO",row["tipo_id"],row["nro_id"],row["nombre"],"NO ENCONTRADO",
                                  empresa_consultada=emp_consultada_m)
                        else:
                            for _,r in dr.iterrows():
                                rd=r.to_dict(); rd["resultado"]="ENCONTRADO EN LISTA"; res.append(rd)
                            log_q("MASIVO",row["tipo_id"],row["nro_id"],row["nombre"],dr["nivel"].iloc[0],
                                  empresa_consultada=emp_consultada_m)
                    prog.empty(); do=pd.DataFrame(res)
                    enc=(do["resultado"]=="ENCONTRADO EN LISTA").sum()
                    co1,co2,co3=st.columns(3)
                    co1.metric("Total",len(di)); co2.metric("🚨 Con coincidencia",enc); co3.metric("✅ Sin coincidencia",len(di)-enc)
                    st.dataframe(do,use_container_width=True,height=280)
                    bx,by=st.columns(2)
                    bx.download_button("📥 Excel",data=a_excel(do),file_name=f"masivo_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
                    if PDF_DISPONIBLE:
                        pm=generar_pdf_masivo(do,umbral_m/100,usuario=st.session_state.user,watermark=not LISTAS_REALES)
                        by.download_button("📄 Reporte PDF",data=pm,file_name=f"reporte_masivo_{date.today()}.pdf",
                            mime="application/pdf",use_container_width=True)
            elif procesar: st.warning("Sube primero un archivo .xlsx")
            else: st.info("Sube un archivo y presiona **Procesar archivo**.")

def mod_policia():
    st.markdown("## 👮 Antecedentes Judiciales — Policía Nacional")
    st.info("Esta fuente usa CAPTCHA. Accede al portal y registra el resultado aquí.", icon="ℹ️")

    uc = st.session_state.ultima_consulta
    if uc["nombre"] or uc["nro_id"]:
        st.success(f"✅ Datos precargados desde última consulta: **{uc['nombre'] or uc['nro_id']}**")

    st.markdown("""<a class="ext-link" href="https://antecedentes.policia.gov.co:7005/WebJudicial/" target="_blank">
      <div class="ext-title">🔗 Policía Nacional — Certificado Judicial</div>
      <div class="ext-url">antecedentes.policia.gov.co</div></a>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 📝 Registrar resultado de consulta manual")
    with st.form("frm_policia", clear_on_submit=True):
        c1, c2 = st.columns(2)
        tid = c1.selectbox("Tipo ID", ["CC","NIT","CE","PAS"],
                           index=["CC","NIT","CE","PAS"].index(uc["tipo_id"]) if uc["tipo_id"] in ["CC","NIT","CE","PAS"] else 0)
        nid = c2.text_input("Número de identificación", value=uc["nro_id"])
        nom = st.text_input("Nombre completo", value=uc["nombre"])
        res = st.radio("Resultado", ["SIN ANTECEDENTES","CON ANTECEDENTES","NO SE PUDO VERIFICAR"], horizontal=True)
        obs = st.text_area("Observaciones", height=80)
        ok  = st.form_submit_button("💾 Guardar registro", type="primary")
    if ok:
        if not nid or not nom: st.warning("Completa todos los campos.")
        else:
            log_q("POLICÍA", tid, nid, nom, res)
            if res == "SIN ANTECEDENTES":    st.success("✅ Sin antecedentes.")
            elif res == "CON ANTECEDENTES":  st.error("🚨 Con antecedentes.")
            else:                            st.warning("⚠️ Verificación pendiente.")
            if PDF_DISPONIBLE:
                pp = generar_pdf_manual(tid, nid, nom, "POLICÍA", res, observacion=obs,
                                        usuario=st.session_state.user, watermark=not LISTAS_REALES)
                st.download_button("📄 Certificado PDF", data=pp,
                                   file_name=f"policia_{nid}_{date.today()}.pdf", mime="application/pdf")

def mod_procuraduria():
    st.markdown("## ⚖️ Antecedentes Disciplinarios — Procuraduría General")
    st.info("Esta fuente usa CAPTCHA. Accede al portal y registra el resultado aquí.", icon="ℹ️")

    uc = st.session_state.ultima_consulta
    if uc["nombre"] or uc["nro_id"]:
        st.success(f"✅ Datos precargados desde última consulta: **{uc['nombre'] or uc['nro_id']}**")

    st.markdown("""<a class="ext-link" href="https://www.procuraduria.gov.co/Pages/Consulta-de-Antecedentes.aspx" target="_blank">
      <div class="ext-title">🔗 Procuraduría General — Antecedentes Disciplinarios</div>
      <div class="ext-url">procuraduria.gov.co</div></a>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 📝 Registrar resultado de consulta manual")
    with st.form("frm_procu", clear_on_submit=True):
        c1, c2 = st.columns(2)
        tid = c1.selectbox("Tipo ID", ["CC","NIT","CE","PAS"],
                           index=["CC","NIT","CE","PAS"].index(uc["tipo_id"]) if uc["tipo_id"] in ["CC","NIT","CE","PAS"] else 0)
        nid = c2.text_input("Número de identificación", value=uc["nro_id"])
        nom = st.text_input("Nombre completo", value=uc["nombre"])
        res = st.radio("Resultado", ["SIN SANCIONES DISCIPLINARIAS","CON SANCIONES DISCIPLINARIAS","NO SE PUDO VERIFICAR"], horizontal=True)
        obs = st.text_area("Observaciones", height=80)
        ok  = st.form_submit_button("💾 Guardar registro", type="primary")
    if ok:
        if not nid or not nom: st.warning("Completa todos los campos.")
        else:
            log_q("PROCURADURÍA", tid, nid, nom, res)
            if "SIN" in res:   st.success("✅ Sin sanciones.")
            elif "CON" in res: st.error("🚨 Con sanciones disciplinarias.")
            else:              st.warning("⚠️ Verificación pendiente.")
            if PDF_DISPONIBLE:
                pp = generar_pdf_manual(tid, nid, nom, "PROCURADURÍA", res, observacion=obs,
                                        usuario=st.session_state.user, watermark=not LISTAS_REALES)
                st.download_button("📄 Certificado PDF", data=pp,
                                   file_name=f"procuraduria_{nid}_{date.today()}.pdf", mime="application/pdf")

def mod_otras_fuentes():
    st.markdown("## 🔗 Otras Fuentes de Consulta")
    st.caption("Accede directamente a cada fuente oficial — los datos de tu última consulta están precargados en los links.")

    uc  = st.session_state.ultima_consulta
    nom = _uparse.quote(uc["nombre"]) if uc["nombre"] else ""
    nid = _uparse.quote(uc["nro_id"]) if uc["nro_id"] else ""

    if uc["nombre"] or uc["nro_id"]:
        st.success(f"✅ Última consulta: **{uc['nombre'] or uc['nro_id']}** — los links incluyen el nombre precargado donde es posible.")

    st.markdown("### 🏛️ Fuentes judiciales y disciplinarias")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"""
        <a class="ext-link" href="https://consultaprocesos.ramajudicial.gov.co/procesos/numeroRadicacion" target="_blank">
          <div class="ext-title">⚖️ Rama Judicial — Consulta de Procesos</div>
          <div class="ext-desc">Consulta de procesos judiciales por nombre, NIT o número de radicación.</div>
          <div class="ext-url">ramajudicial.gov.co</div>
        </a>
        <a class="ext-link" href="https://www.fiscalia.gov.co/colombia/?s={nom}" target="_blank">
          <div class="ext-title">🏛️ Fiscalía General de la Nación</div>
          <div class="ext-desc">Búsqueda de boletines, capturas, imputados y condenas.</div>
          <div class="ext-url">fiscalia.gov.co/?s={uc['nombre'] or '(nombre)'}</div>
        </a>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <a class="ext-link" href="https://www.procuraduria.gov.co/Pages/Consulta-de-Antecedentes.aspx" target="_blank">
          <div class="ext-title">📋 Procuraduría — Antecedentes Disciplinarios</div>
          <div class="ext-desc">Sanciones disciplinarias de servidores públicos y particulares.</div>
          <div class="ext-url">procuraduria.gov.co</div>
        </a>
        <a class="ext-link" href="https://antecedentes.policia.gov.co:7005/WebJudicial/" target="_blank">
          <div class="ext-title">👮 Policía Nacional — Certificado Judicial</div>
          <div class="ext-desc">Antecedentes judiciales de personas naturales.</div>
          <div class="ext-url">antecedentes.policia.gov.co</div>
        </a>
        """, unsafe_allow_html=True)

    st.markdown("### 🏢 Fuentes empresariales y registros")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"""
        <a class="ext-link" href="https://ruesfront.rues.org.co/" target="_blank">
          <div class="ext-title">🏢 RUES — Registro Único Empresarial</div>
          <div class="ext-desc">Consulta de empresas, representantes legales y estados societarios en Colombia.</div>
          <div class="ext-url">rues.org.co</div>
        </a>
        <a class="ext-link" href="https://ruaf.sispro.gov.co/Filtro.aspx" target="_blank">
          <div class="ext-title">📋 RUAF — Registro Único de Afiliados</div>
          <div class="ext-desc">Verificación de afiliación a seguridad social y empresas ficticias.</div>
          <div class="ext-url">ruaf.sispro.gov.co</div>
        </a>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <a class="ext-link" href="https://www.contraloria.gov.co/contraloria/consultasRegistro/consultaBoletinResponsables" target="_blank">
          <div class="ext-title">🔍 Contraloría — Boletín de Responsables</div>
          <div class="ext-desc">Personas con responsabilidad fiscal — inhabilitadas para contratos públicos.</div>
          <div class="ext-url">contraloria.gov.co</div>
        </a>
        <a class="ext-link" href="https://www1.funcionpublica.gov.co/fdci/consultaCiudadana/consultaPEP" target="_blank">
          <div class="ext-title">🎖️ Función Pública — PEPs Colombia</div>
          <div class="ext-desc">Lista oficial de Personas Expuestas Políticamente (Decreto 830/2021).</div>
          <div class="ext-url">funcionpublica.gov.co</div>
        </a>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📝 Registrar resultado de consulta en fuente externa")
    with st.form("frm_otras", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        fuente_sel = c1.selectbox("Fuente consultada", [
            "Rama Judicial","RUES","RUAF","Contraloría","Función Pública / PEPs","Otra"
        ])
        tid = c2.selectbox("Tipo ID", ["CC","NIT","CE","PAS"],
                           index=["CC","NIT","CE","PAS"].index(uc["tipo_id"]) if uc["tipo_id"] in ["CC","NIT","CE","PAS"] else 0)
        nid = c3.text_input("Número de identificación", value=uc["nro_id"])
        nom = st.text_input("Nombre completo", value=uc["nombre"])
        res = st.radio("Resultado", ["SIN HALLAZGOS","CON HALLAZGOS","NO SE PUDO VERIFICAR"], horizontal=True)
        obs = st.text_area("Observaciones / detalle del hallazgo", height=80)
        ok  = st.form_submit_button("💾 Guardar registro", type="primary")
    if ok:
        if not nid or not nom: st.warning("Completa todos los campos.")
        else:
            log_q(fuente_sel.upper(), tid, nid, nom, res)
            if "SIN" in res:   st.success(f"✅ {fuente_sel} — Sin hallazgos.")
            elif "CON" in res: st.error(f"🚨 {fuente_sel} — Con hallazgos. Escalar al Oficial de Cumplimiento.")
            else:              st.warning(f"⚠️ {fuente_sel} — Verificación pendiente.")
            if PDF_DISPONIBLE:
                pp = generar_pdf_manual(tid, nid, nom, fuente_sel, res,
                                        observacion=obs, usuario=st.session_state.user,
                                        watermark=not LISTAS_REALES)
                nombre_arch = fuente_sel.lower().replace(" ","_").replace("/","_")
                st.download_button("📄 Certificado PDF", data=pp,
                                   file_name=f"{nombre_arch}_{nid}_{date.today()}.pdf",
                                   mime="application/pdf")

def mod_logs():
    st.markdown("## 📋 Registros de consultas")
    info = st.session_state.get("user_info", {})
    es_super = info.get("rol") == "superadmin"
    empresa_id = None if es_super else info.get("empresa_id")
    rows = listar_consultas(empresa_id=empresa_id)
    if not rows:
        st.info("No hay registros aún."); return
    df = pd.DataFrame(rows)
    c1, c2, c3 = st.columns(3)
    fm = c1.selectbox("Módulo",  ["Todos"] + sorted(df["modulo"].dropna().unique().tolist()))
    fu = c2.selectbox("Usuario", ["Todos"] + sorted(df["username"].dropna().unique().tolist()))
    ft = c3.text_input("Buscar nombre o número")
    dff = df.copy()
    if fm != "Todos": dff = dff[dff["modulo"] == fm]
    if fu != "Todos": dff = dff[dff["username"] == fu]
    if ft:
        q = ft.upper()
        dff = dff[dff["nombre"].str.upper().str.contains(q, na=False) |
                  dff["nro_id"].str.contains(q, na=False)]
    dff = dff.sort_values("fecha_hora", ascending=False).reset_index(drop=True)
    cols_show = ["fecha_hora","username","empresa_consultada","modulo","tipo_id","nro_id","nombre","resultado"]
    cols_show = [c for c in cols_show if c in dff.columns]
    st.caption(f"Mostrando {len(dff)} de {len(df)} registros")
    st.dataframe(dff[cols_show], use_container_width=True, height=400)
    st.download_button("📥 Exportar", data=a_excel(dff[cols_show]),
        file_name=f"logs_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def mod_stats():
    st.markdown("## 📊 Estadísticas de uso")
    info = st.session_state.get("user_info", {})
    es_super = info.get("rol") == "superadmin"
    empresa_id = None if es_super else info.get("empresa_id")
    rows = listar_consultas(empresa_id=empresa_id)
    if not rows:
        st.info("No hay registros de consultas aún."); return
    df_logs = pd.DataFrame(rows)
    df_logs["fecha_hora"] = pd.to_datetime(df_logs["fecha_hora"], format="mixed", dayfirst=False)

    mes_actual = date.today().replace(day=1)
    df_mes = df_logs[df_logs["fecha_hora"].dt.date >= mes_actual]

    ALERTAS = {"EXACTA","APROXIMADA","CON ANTECEDENTES","CON SANCIONES DISCIPLINARIAS","CON HALLAZGOS"}
    def es_alerta(r): return any(a in str(r).upper() for a in ALERTAS) or "NOTICIAS ADVERSAS" in str(r).upper()

    total_mes  = len(df_mes)
    alertas    = int(df_logs["resultado"].apply(es_alerta).sum())
    usuarios_a = int(df_logs["username"].nunique())
    masivos    = int((df_logs["modulo"] == "MASIVO").sum())

    c1,c2,c3,c4=st.columns(4)
    c1.markdown(f'<div class="metric-box"><div class="metric-num">{total_mes}</div><div class="metric-label">Consultas este mes</div></div>',unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-box"><div class="metric-num">{alertas}</div><div class="metric-label">🚨 Alertas generadas</div></div>',unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-box"><div class="metric-num">{usuarios_a}</div><div class="metric-label">Usuarios activos</div></div>',unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-box"><div class="metric-num">{masivos}</div><div class="metric-label">Cargas masivas</div></div>',unsafe_allow_html=True)
    st.markdown("---")

    co1,co2=st.columns(2)
    with co1:
        st.markdown("#### Consultas por módulo")
        por_mod = df_logs.groupby("modulo").size().rename("Consultas")
        st.bar_chart(por_mod, height=220)
    with co2:
        st.markdown("#### Distribución de resultados")
        df_logs["tipo_resultado"] = df_logs["resultado"].apply(
            lambda r: "Alerta" if es_alerta(r) else "Sin coincidencia")
        dist = df_logs["tipo_resultado"].value_counts().rename("Cantidad")
        st.bar_chart(dist, height=220)

    st.markdown("---")
    co1,co2=st.columns(2)
    with co1:
        st.markdown("#### Por usuario")
        por_usr = (df_logs.groupby("username")
                   .agg(Consultas=("resultado","count"),
                        Alertas=("resultado", lambda x: x.apply(es_alerta).sum()))
                   .reset_index().rename(columns={"username":"Usuario"}))
        st.dataframe(por_usr, use_container_width=True, hide_index=True)
    with co2:
        st.markdown("#### Últimas 10 consultas")
        st.dataframe(
            df_logs.sort_values("fecha_hora", ascending=False)
                   .head(10)[["fecha_hora","username","modulo","nombre","resultado"]]
                   .rename(columns={"fecha_hora":"Fecha","username":"Usuario",
                                    "modulo":"Módulo","nombre":"Nombre","resultado":"Resultado"}),
            use_container_width=True, hide_index=True)
    st.markdown("---")
    st.download_button("📥 Exportar registros", data=a_excel(df_logs.astype(str)),
        file_name=f"estadisticas_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def mod_empresas():
    st.markdown("## 🏢 Gestión de Empresas")
    empresas = listar_empresas()

    with st.expander("➕ Registrar nueva empresa", expanded=False):
        with st.form("frm_nueva_empresa", clear_on_submit=True):
            c1, c2 = st.columns(2)
            nom_e = c1.text_input("Nombre de la empresa *")
            nit_e = c2.text_input("NIT")
            ok_e = st.form_submit_button("💾 Crear empresa", type="primary")
        if ok_e:
            if not nom_e.strip():
                st.warning("El nombre de la empresa es obligatorio.")
            else:
                crear_empresa(nom_e, nit_e)
                st.success(f"✅ Empresa '{nom_e}' creada."); st.rerun()

    st.markdown("---")
    st.markdown(f"**{len(empresas)} empresa(s) registradas**")
    for emp in empresas:
        activo = bool(emp["activo"])
        badge = "🟢" if activo else "🔴"
        with st.expander(f"{badge} {emp['nombre']}  —  NIT: {emp['nit'] or '—'}", expanded=False):
            c1, c2, c3 = st.columns(3)
            c1.caption(f"ID: {emp['id']}")
            c2.caption(f"Registrada: {emp['fecha_registro'][:10] if emp.get('fecha_registro') else '—'}")
            lbl = "⛔ Desactivar" if activo else "✅ Activar"
            if c3.button(lbl, key=f"tog_emp_{emp['id']}"):
                toggle_empresa(emp["id"], not activo)
                st.rerun()


def mod_usuarios():
    st.markdown("## 👥 Gestión de Usuarios")
    empresas = listar_empresas()
    empresas_activas = [e for e in empresas if e["activo"]]
    emp_map = {e["nombre"]: e["id"] for e in empresas_activas}

    with st.expander("➕ Crear nuevo analista", expanded=False):
        with st.form("frm_nuevo_usuario", clear_on_submit=True):
            c1, c2 = st.columns(2)
            emp_sel = c1.selectbox("Empresa *", list(emp_map.keys()))
            username_n = c2.text_input("Username *", placeholder="sin espacios")
            c3, c4 = st.columns(2)
            nombre_n = c3.text_input("Nombre completo *")
            pwd_n = c4.text_input("Contraseña inicial *", type="password")
            ok_u = st.form_submit_button("💾 Crear usuario", type="primary")
        if ok_u:
            if not all([emp_sel, username_n.strip(), nombre_n.strip(), pwd_n]):
                st.warning("Completa todos los campos.")
            else:
                try:
                    crear_usuario(username_n.strip(), pwd_n, nombre_n.strip(), emp_map[emp_sel])
                    st.success(f"✅ Usuario '{username_n}' creado para {emp_sel}."); st.rerun()
                except Exception as ex:
                    st.error(f"Error al crear usuario: {ex}")

    st.markdown("---")
    emp_filtro_opts = ["Todas"] + [e["nombre"] for e in empresas]
    filtro_emp = st.selectbox("Filtrar por empresa", emp_filtro_opts)
    emp_id_filtro = None if filtro_emp == "Todas" else next(
        (e["id"] for e in empresas if e["nombre"] == filtro_emp), None)
    usuarios = listar_usuarios(empresa_id=emp_id_filtro)

    st.markdown(f"**{len(usuarios)} usuario(s)**")
    for u in usuarios:
        activo = bool(u["activo"])
        badge = "🟢" if activo else "🔴"
        rol_label = "Superadmin" if u["rol"] == "superadmin" else "Analista"
        with st.expander(f"{badge} {u['nombre']}  ({u['username']})  —  {u.get('empresa_nombre','—')}  |  {rol_label}", expanded=False):
            c1, c2, c3 = st.columns(3)
            c1.caption(f"Creado: {u['fecha_creacion'][:10] if u.get('fecha_creacion') else '—'}")
            lbl = "⛔ Desactivar" if activo else "✅ Activar"
            if u["rol"] != "superadmin":
                if c2.button(lbl, key=f"tog_usr_{u['id']}"):
                    toggle_usuario(u["id"], not activo)
                    st.rerun()
            with c3.expander("🔑 Reset contraseña"):
                with st.form(f"frm_reset_{u['id']}", clear_on_submit=True):
                    nueva_pwd = st.text_input("Nueva contraseña", type="password", key=f"npwd_{u['id']}")
                    if st.form_submit_button("Cambiar"):
                        if nueva_pwd:
                            reset_password(u["id"], nueva_pwd)
                            st.success("Contraseña actualizada.")

def mod_perfil():
    st.markdown("## 🔑 Mi perfil")
    info = st.session_state.get("user_info", {})

    c1, c2, c3 = st.columns(3)
    c1.metric("Usuario", info.get("username", st.session_state.user))
    c2.metric("Nombre", info.get("nombre", "—"))
    c3.metric("Empresa", info.get("empresa_nombre", "—"))
    st.markdown("---")
    st.markdown("### Cambiar contraseña")

    with st.form("frm_cambiar_pwd", clear_on_submit=True):
        pwd_actual  = st.text_input("Contraseña actual", type="password")
        pwd_nueva   = st.text_input("Nueva contraseña", type="password")
        pwd_confirm = st.text_input("Confirmar nueva contraseña", type="password")
        ok = st.form_submit_button("💾 Cambiar contraseña", type="primary")

    if ok:
        if not pwd_actual or not pwd_nueva or not pwd_confirm:
            st.warning("Completa todos los campos.")
        elif pwd_nueva != pwd_confirm:
            st.error("La nueva contraseña y la confirmación no coinciden.")
        elif len(pwd_nueva) < 6:
            st.error("La contraseña debe tener al menos 6 caracteres.")
        else:
            verificado = verificar_usuario(st.session_state.user, pwd_actual)
            if not verificado:
                st.error("La contraseña actual es incorrecta.")
            else:
                reset_password(info["id"], pwd_nueva)
                st.success("✅ Contraseña actualizada correctamente.")


if not st.session_state.logged_in:
    pantalla_login()
else:
    sidebar()
    m = st.session_state.menu
    if   "Unificada"                       in m: mod_unificada()
    elif "Otras"                           in m: mod_otras_fuentes()
    elif "Registros"                       in m: mod_logs()
    elif "Estadisticas" in m or "Estadísticas" in m: mod_stats()
    elif "Empresas"                        in m: mod_empresas()
    elif "Usuarios"                        in m: mod_usuarios()
    elif "perfil"    in m or "Perfil"      in m: mod_perfil()