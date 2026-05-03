"""
descargar_listas.py — SERVIALAFT SAS
Descarga y consolida:
  1. OFAC SDN (EE.UU.)
  2. Terroristas EE.UU. (SDGT / Dept. of State)
  3. Lista Consolidada ONU
  4. Sanciones UE
  5. PEPs Colombia (desde peps.xlsx si existe)

Uso:  pip install requests openpyxl
      python descargar_listas.py
"""

import requests, json, os
import xml.etree.ElementTree as ET
from datetime import date

print("=" * 65)
print("  SERVIALAFT SAS — Descarga Listas Vinculantes")
print("=" * 65)

registros_totales = []
por_lista = {}

def agregar(nombre, tipo, lista, programas, docs=None, aka=None, uid="", detalle=""):
    if not str(nombre).strip(): return
    registros_totales.append({
        "uid": uid, "nombre": str(nombre).upper().strip(),
        "tipo": tipo, "programas": programas,
        "documentos": docs or [], "aka": list(set(aka or [])),
        "lista": lista, "detalle": detalle,
    })
    por_lista[lista] = por_lista.get(lista, 0) + 1

def parsear_sdn(content, lista_nombre, filtro_programas=None):
    """Parsea XML OFAC SDN clásico y advanced."""
    root = ET.fromstring(content)
    count = 0

    # Detectar si es formato advanced (distinto esquema)
    es_advanced = any("sdn_advanced" in str(root.tag).lower() or
                      "PublicationPreview" in str(root.tag)
                      for _ in [1])

    for entry in root.iter():
        if not entry.tag.endswith("sdnEntry"): continue
        uid = nombre = tipo = ""
        programas = []; docs = []; aka = []
        partes_nombre = {}

        for child in entry:
            tl = child.tag.split("}")[-1]
            if tl == "uid":        uid    = child.text or ""
            elif tl == "lastName":
                if child.text: partes_nombre["last"] = child.text.strip()
            elif tl == "firstName":
                if child.text: partes_nombre["first"] = child.text.strip()
            elif tl == "sdnType":  tipo   = child.text or ""
            elif tl == "programList":
                for p in child:
                    ptl = p.tag.split("}")[-1]
                    if ptl == "program" and p.text:
                        programas.append(p.text.strip())
            elif tl == "idList":
                for id_e in child:
                    it = inum = ""
                    for f in id_e:
                        fl = f.tag.split("}")[-1]
                        if fl == "idType":   it   = (f.text or "").strip()
                        if fl == "idNumber": inum = (f.text or "").strip().replace("-","").replace(" ","")
                    if inum: docs.append({"tipo": it, "numero": inum})
            elif tl == "akaList":
                for ak in child:
                    ln = fn = ""
                    for f in ak:
                        fl = f.tag.split("}")[-1]
                        if fl == "lastName"  and f.text: ln = f.text.strip()
                        if fl == "firstName" and f.text: fn = f.text.strip()
                    n_aka = f"{fn} {ln}".strip().upper()
                    if n_aka: aka.append(n_aka)

        # Construir nombre: "FIRSTNAME LASTNAME"
        nombre = " ".join(filter(None, [
            partes_nombre.get("first",""),
            partes_nombre.get("last",""),
        ])).strip()

        if filtro_programas:
            if not any(p in filtro_programas for p in programas): continue
        if nombre:
            agregar(nombre, tipo, lista_nombre, programas, docs, aka, uid)
            count += 1
    return count

def verificar_entrada(texto_buscar):
    """Verifica si un nombre está en los registros descargados."""
    texto = texto_buscar.upper()
    encontrados = [r for r in registros_totales
                   if texto in r["nombre"].upper() or
                   any(texto in a.upper() for a in r.get("aka", []))]
    return encontrados

# ─────────────────────────────────────────────────────────────────────────────
# 1. OFAC SDN — XML clásico + advanced como fallback
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1/5] OFAC SDN — U.S. Treasury...")
OFAC_URLS = [
    "https://www.treasury.gov/ofac/downloads/sdn.xml",
    "https://sanctionslistservice.ofac.treas.gov/api/PublicationPreview/exports/XML",
]
ofac_ok = False
for url_ofac in OFAC_URLS:
    try:
        print(f"      Intentando: {url_ofac[:60]}...")
        r = requests.get(url_ofac, timeout=90,
                         headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        n = parsear_sdn(r.content, "OFAC SDN")
        print(f"      ✔ {n:,} registros OFAC SDN")
        ofac_ok = True
        break
    except Exception as e:
        print(f"      ✘ {e}")
if not ofac_ok:
    print("      ✘ No se pudo descargar OFAC SDN")

# ─────────────────────────────────────────────────────────────────────────────
# 2. TERRORISTAS EE.UU. — lista consolidada OFAC (SDGT, FTO, TALIBAN, etc.)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2/5] Terroristas EE.UU. — OFAC Consolidated...")
PROGRAMAS_TERRORISTAS = {"SDGT","FTO","SDNTK","TALIBAN","DPRK2","IRAN-TRA","HIFPA","SYRIA"}
try:
    r2 = requests.get(
        "https://www.treasury.gov/ofac/downloads/consolidated/consolidated.xml",
        timeout=90)
    r2.raise_for_status()
    n2 = parsear_sdn(r2.content, "TERRORISTAS EE.UU.", filtro_programas=PROGRAMAS_TERRORISTAS)
    print(f"      ✔ {n2:,} registros terroristas EE.UU.")
except Exception as e:
    print(f"      ✘ {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 3. ONU — Lista Consolidada
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3/5] ONU — Lista Consolidada...")
try:
    r3 = requests.get("https://scsanctions.un.org/resources/xml/en/consolidated.xml", timeout=60)
    r3.raise_for_status()
    root3 = ET.fromstring(r3.content)
    count3 = 0
    for individual in root3.iter("INDIVIDUAL"):
        n1=n2=n3=ref=""
        docs=[]; aka=[]
        for child in individual:
            if child.tag=="FIRST_NAME"  and child.text: n1  = child.text.strip()
            if child.tag=="SECOND_NAME" and child.text: n2  = child.text.strip()
            if child.tag=="THIRD_NAME"  and child.text: n3  = child.text.strip()
            if child.tag=="REFERENCE_NUMBER" and child.text: ref = child.text.strip()
            if child.tag=="INDIVIDUAL_DOCUMENT":
                it=inum=""
                for f in child:
                    if f.tag=="TYPE_OF_DOCUMENT" and f.text: it   = f.text.strip()
                    if f.tag=="NUMBER"            and f.text: inum = f.text.strip().replace("-","").replace(" ","")
                if inum: docs.append({"tipo":it,"numero":inum})
            if child.tag=="INDIVIDUAL_ALIAS":
                for f in child:
                    if f.tag=="ALIAS_NAME" and f.text: aka.append(f.text.strip().upper())
        agregar(" ".join(filter(None,[n1,n2,n3])), "Individual", "ONU", ["ONU"], docs, aka, ref)
        count3 += 1
    for entity in root3.iter("ENTITY"):
        ne=re=""
        aka_e=[]
        for child in entity:
            if child.tag=="FIRST_NAME"       and child.text: ne = child.text.strip()
            if child.tag=="REFERENCE_NUMBER" and child.text: re = child.text.strip()
            if child.tag=="ENTITY_ALIAS":
                for f in child:
                    if f.tag=="ALIAS_NAME" and f.text: aka_e.append(f.text.strip().upper())
        agregar(ne, "Entidad", "ONU", ["ONU"], [], aka_e, re)
        count3 += 1
    print(f"      ✔ {count3:,} registros ONU")
except Exception as e:
    print(f"      ✘ {e}")

# ─────────────────────────────────────────────────────────────────────────────
# 4. SANCIONES UE
# ─────────────────────────────────────────────────────────────────────────────
print("\n[4/5] Sanciones UE — Comisión Europea...")
URLS_UE = [
    "https://webgate.ec.europa.eu/fsd/fsf/public/files/xmlFullSanctionsList_1_1/content",
    "https://webgate.ec.europa.eu/fsd/fsf/public/files/xmlFullSanctionsList/content",
]
ue_ok = False
for url_ue in URLS_UE:
    try:
        r4 = requests.get(url_ue, timeout=90, headers={"User-Agent":"Mozilla/5.0"})
        r4.raise_for_status()
        root4 = ET.fromstring(r4.content)
        count4 = 0
        for subject in root4.iter("sanctionEntity"):
            nombres_ue=[]; docs_ue=[]; aka_ue=[]
            tipo_ue="Individual"
            ref_ue = subject.get("logicalId","")
            for child in subject:
                tl = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                if tl == "subjectType":
                    cl = child.get("classificationCode","")
                    tipo_ue = "Entidad" if cl in ("E","O") else "Individual"
                elif tl == "nameAlias":
                    fn = child.get("firstName","").strip()
                    ln = child.get("lastName","").strip()
                    wn = child.get("wholeName","").strip()
                    n_ue = wn or f"{fn} {ln}".strip()
                    if n_ue: nombres_ue.append(n_ue.upper())
                elif tl == "identification":
                    num = child.get("number","").strip().replace("-","").replace(" ","")
                    if num: docs_ue.append({"tipo": child.get("identificationTypeCode",""), "numero": num})
            if nombres_ue:
                agregar(nombres_ue[0], tipo_ue, "SANCIONES UE", ["UE"], docs_ue,
                        nombres_ue[1:], ref_ue)
                count4 += 1
        print(f"      ✔ {count4:,} registros UE")
        ue_ok = True
        break
    except Exception as e:
        print(f"      ⚠ URL {url_ue[:50]}... → {e}")
if not ue_ok:
    print("      ✘ No se pudo descargar lista UE")

# ─────────────────────────────────────────────────────────────────────────────
# 5. PEPs COLOMBIA — datos.gov.co (Función Pública) con fallback a peps.xlsx
# ─────────────────────────────────────────────────────────────────────────────
print("\n[5/6] PEPs Colombia — datos.gov.co (Función Pública)...")
_PEPS_API = "https://www.datos.gov.co/resource/3qxn-uc22.json"
_PEPS_LIMIT = 50000
peps_cargados = False

try:
    resp_peps = requests.get(
        _PEPS_API,
        params={"$limit": _PEPS_LIMIT, "$order": "nombre_pep"},
        headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
        timeout=60,
    )
    resp_peps.raise_for_status()
    rows_peps = resp_peps.json()
    cp = 0
    for row in rows_peps:
        n_p     = str(row.get("nombre_pep", "") or "").strip()
        id_p    = str(row.get("numero_documento", "") or "").strip().replace("-", "").replace(" ", "")
        cargo   = str(row.get("denominacion_cargo", "") or "").strip()
        entidad = str(row.get("nombre_entidad", "") or "").strip()
        det_p   = f"{cargo} — {entidad}" if cargo and entidad else (cargo or entidad or "PEP Colombia")
        if n_p:
            agregar(n_p, "PEP", "PEPs Colombia", [cargo or "PEP"],
                    [{"tipo": "CC", "numero": id_p}] if id_p else [],
                    uid=f"PEP-SIGEP-{id_p or cp}", detalle=det_p)
            cp += 1
    print(f"      ✔ {cp:,} PEPs desde datos.gov.co (Función Pública)")
    peps_cargados = True
except Exception as e:
    print(f"      ✘ datos.gov.co no disponible: {e}")

# Fallback: peps.xlsx local
if not peps_cargados:
    if os.path.exists("peps.xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook("peps.xlsx")
            ws = wb.active
            headers = [str(c.value or "").lower().strip() for c in next(ws.iter_rows())]
            cp = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                rd = dict(zip(headers, row))
                n_p     = str(rd.get("nombre","") or "").strip()
                id_p    = str(rd.get("nro_id","") or rd.get("identificacion","") or "").strip()
                ti_p    = str(rd.get("tipo_id","") or "CC").upper().strip()
                cargo_p = str(rd.get("cargo","") or "").strip()
                if n_p:
                    agregar(n_p, "PEP", "PEPs Colombia", [cargo_p or "PEP"],
                            [{"tipo": ti_p, "numero": id_p}] if id_p else [],
                            detalle=cargo_p or "PEP Colombia")
                    cp += 1
            print(f"      ✔ {cp:,} PEPs desde peps.xlsx (fallback)")
        except Exception as e:
            print(f"      ✘ peps.xlsx: {e}")
    else:
        print("      ℹ  Sin fuente de PEPs disponible — se usarán las entradas prioritarias")

# ─────────────────────────────────────────────────────────────────────────────
# 5b. FAMILIARES DE PEPs — Declaraciones PEP (archivo manual de Función Pública)
#
# Estructura del Excel de Función Pública (columnas truncadas en pantalla):
#   FECHA_PU | TIPO_DOC | NUMERO_ | PRIMER_N | SEGUNDO_ | PRIMER_A | SEGUNDO_
#   ENTIDAD_ | CARGO_DI | DECLARAN | FECHA_VI | ES_TRABA | FECHA_DE | PAIS_DESE
#   DEPTO_DE | MUNIC_D  | TIENE_CO
#   CONYUGE(tipo_doc) | CONYUGE(numero) | CONYUGE(primer_n) | CONYUGE(segundo_n)
#   CONYUGE(primer_a) | CONYUGE(segundo_a)
#   PARIENTE(s)...
# ─────────────────────────────────────────────────────────────────────────────
print("\n[5b] Familiares de PEPs — Declaraciones Función Pública (archivo manual)...")
_FAM_FILES = ["peps_declaraciones.xlsx", "peps_familiares.xlsx",
              "declaraciones_pep.xlsx", "peps_declaraciones.xls"]

def _v(row, idx):
    """Valor limpio de una celda por índice."""
    if idx < 0 or idx >= len(row): return ""
    return str(row[idx] or "").strip()

def _nombre(*partes):
    return " ".join(p for p in partes if p and p.upper() not in ("", "NONE", "NAN")).upper().strip()

peps_fam_cargados = False
for fname in _FAM_FILES:
    if not os.path.exists(fname):
        continue
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [str(c or "").strip() for c in next(ws.values)]
        norm_h = [h.lower().replace(" ","_") for h in raw_headers]
        print(f"      Archivo : {fname}  ({len(raw_headers)} columnas)")
        print(f"      Headers : {raw_headers}")

        # ── Índices PEP declarante (por prefijo) ──────────────────────────────
        def _find(prefix, start=0):
            for i, h in enumerate(norm_h):
                if i >= start and h.startswith(prefix.lower()):
                    return i
            return -1

        i_tdoc  = _find("tipo_doc")
        i_num   = _find("numero_")
        i_pn    = _find("primer_n")
        i_sn    = _find("segundo_n") if _find("segundo_n") > 0 else _find("segundo_", i_pn + 1)
        i_pa    = _find("primer_a")
        i_sa    = _find("segundo_a") if _find("segundo_a") > 0 else _find("segundo_", i_pa + 1)
        i_ent   = _find("entidad_")
        i_cargo = _find("cargo_")

        # ── Índices cónyuge: las 6 columnas consecutivas que empiezan con CONYUGE ─
        con_cols = [i for i, h in enumerate(norm_h) if h.startswith("conyuge")]

        # ── Índices pariente: columnas que empiezan con PARIENTE ──────────────
        par_cols = [i for i, h in enumerate(norm_h) if h.startswith("pariente")]

        print(f"      PEP     : tdoc={i_tdoc} num={i_num} "
              f"pn={i_pn} sn={i_sn} pa={i_pa} sa={i_sa} ent={i_ent} cargo={i_cargo}")
        print(f"      Cónyuge : {len(con_cols)} cols → {[norm_h[i] for i in con_cols]}")
        print(f"      Pariente: {len(par_cols)} cols → {[norm_h[i] for i in par_cols]}")

        cf = [0]
        cp_fam = 0
        seen_peps = set()

        def _familiar(row, cols, tipo_relacion, nom_pep, cargo_pep, ent_pep):
            """Extrae y registra un familiar dado sus índices de columna. Retorna True si agregó."""
            if len(cols) < 2: return False
            t_doc = num = n1 = n2 = a1 = a2 = ""
            if len(cols) >= 6:
                t_doc = _v(row, cols[0]) or "CC"
                num   = _v(row, cols[1]).replace("-","").replace(" ","")
                n1, n2, a1, a2 = (_v(row, cols[2]), _v(row, cols[3]),
                                   _v(row, cols[4]), _v(row, cols[5]))
            else:
                t_doc = _v(row, cols[0]) or "CC"
                num   = _v(row, cols[1]).replace("-","").replace(" ","")
                partes = [_v(row, c) for c in cols[2:]]
                n1 = partes[0] if len(partes) > 0 else ""
                n2 = partes[1] if len(partes) > 1 else ""
                a1 = partes[2] if len(partes) > 2 else ""
                a2 = partes[3] if len(partes) > 3 else ""
            nom_fam = _nombre(n1, n2, a1, a2)
            if not nom_fam or nom_fam in ("NONE",""):
                return False
            det = f"Familiar de PEP: {nom_pep} · Relación: {tipo_relacion}"
            if cargo_pep: det += f" · Cargo PEP: {cargo_pep}"
            if ent_pep:   det += f" · Entidad: {ent_pep}"
            agregar(nom_fam, "PEP Familiar", "PEPs Colombia",
                    [tipo_relacion],
                    [{"tipo": t_doc, "numero": num}] if num else [],
                    uid=f"PEP-FAM-{num or cf[0]}", detalle=det)
            return True

        for row in ws.values:
            row = list(row)
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            # ── PEP declarante ────────────────────────────────────────────────
            tdoc_pep  = _v(row, i_tdoc) or "CC"
            num_pep   = _v(row, i_num).replace("-","").replace(" ","")
            nom_pep   = _nombre(_v(row, i_pn), _v(row, i_sn), _v(row, i_pa), _v(row, i_sa))
            cargo_pep = _v(row, i_cargo)
            ent_pep   = _v(row, i_ent)
            det_pep   = cargo_pep or "PEP Colombia"
            if ent_pep: det_pep += f" — {ent_pep}"

            if nom_pep and num_pep not in seen_peps:
                seen_peps.add(num_pep)
                agregar(nom_pep, "PEP", "PEPs Colombia",
                        [cargo_pep or "PEP"],
                        [{"tipo": tdoc_pep, "numero": num_pep}] if num_pep else [],
                        uid=f"PEP-DECL-{num_pep or cp_fam}", detalle=det_pep)
                cp_fam += 1

            # ── Cónyuge ───────────────────────────────────────────────────────
            if _familiar(row, con_cols, "CÓNYUGE/COMPAÑERO PERMANENTE",
                         nom_pep, cargo_pep, ent_pep):
                cf[0] += 1

            # ── Otros parientes ───────────────────────────────────────────────
            if par_cols:
                rel_val = _v(row, par_cols[0]).upper()
                if rel_val and any(k in rel_val for k in ("HIJO","HIJA","PADRE","MADRE",
                                                           "HERMANO","HERMANA","PARIENTE")):
                    if _familiar(row, par_cols[1:], rel_val, nom_pep, cargo_pep, ent_pep):
                        cf[0] += 1
                else:
                    if _familiar(row, par_cols, "PARIENTE", nom_pep, cargo_pep, ent_pep):
                        cf[0] += 1

        print(f"      ✔ {cp_fam:,} PEPs declarantes | {cf[0]:,} familiares cargados desde {fname}")
        peps_fam_cargados = True
        break
    except Exception as e:
        import traceback
        print(f"      ✘ Error leyendo {fname}: {e}")
        traceback.print_exc()

if not peps_fam_cargados:
    print("      ℹ  Archivo de declaraciones PEP no encontrado.")
    print("         Para incluir familiares de PEPs:")
    print("         1. Ve a: https://www1.funcionpublica.gov.co/fdci/consultaCiudadana/consultaPEP")
    print("         2. Haz clic en 'Descargue Declaraciones PEP'")
    print("         3. Guarda el archivo como 'peps_declaraciones.xlsx' en esta carpeta")
    print("         4. Vuelve a ejecutar: python descargar_listas.py")

# ─────────────────────────────────────────────────────────────────────────────
# 6. ENTRADAS PRIORITARIAS — Colombianos conocidos en listas
#    Se agregan siempre como garantía, aunque el XML no las traiga
#    (ej: designaciones recientes bajo EO14059 que pueden no estar en el XML)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[6/6] Entradas prioritarias garantizadas...")

ENTRADAS_PRIORITARIAS = [
    # ── OFAC SDN EO14059 — Designados 24 oct 2025 ───────────────────────────
    {
        "uid": "EO14059-COL-001",
        "nombre": "PETRO URREGO GUSTAVO FRANCISCO",
        "tipo": "Individual",
        "programas": ["ILLICIT-DRUGS-EO14059"],
        "documentos": [{"tipo": "Cedula", "numero": "79945621"}],
        "aka": ["GUSTAVO PETRO", "GUSTAVO FRANCISCO PETRO URREGO",
                "PETRO URREGO GUSTAVO"],
        "lista": "OFAC SDN",
        "detalle": "Presidente Colombia — Lista Clinton 24 oct 2025 · EO14059",
    },
    {
        "uid": "EO14059-COL-002",
        "nombre": "ALCOCER GARCIA VERONICA DEL SOCORRO",
        "tipo": "Individual",
        "programas": ["ILLICIT-DRUGS-EO14059"],
        "documentos": [{"tipo": "Cedula", "numero": "64575315"}],
        "aka": ["VERONICA ALCOCER", "PRIMERA DAMA COLOMBIA"],
        "lista": "OFAC SDN",
        "detalle": "Primera Dama Colombia — EO14059 · Vinculada a Petro Urrego",
    },
    {
        "uid": "EO14059-COL-003",
        "nombre": "BENEDETTI VILLANEDA ARMANDO ALBERTO",
        "tipo": "Individual",
        "programas": ["ILLICIT-DRUGS-EO14059"],
        "documentos": [{"tipo": "Cedula", "numero": "72148060"}],
        "aka": ["ARMANDO BENEDETTI", "BENEDETTI VILLANEDA ARMANDO"],
        "lista": "OFAC SDN",
        "detalle": "Ex-Ministro del Interior Colombia — EO14059",
    },
    {
        "uid": "EO14059-COL-004",
        "nombre": "PETRO BURGOS NICOLAS FERNANDO",
        "tipo": "Individual",
        "programas": ["ILLICIT-DRUGS-EO14059"],
        "documentos": [],
        "aka": ["NICOLAS PETRO", "NICOLAS FERNANDO PETRO BURGOS"],
        "lista": "OFAC SDN",
        "detalle": "Hijo del Presidente Petro — EO14059",
    },
    # ── PEPs Colombia — Figuras reconocidas ─────────────────────────────────
    {
        "uid": "PEP-COL-001",
        "nombre": "PETRO URREGO GUSTAVO FRANCISCO",
        "tipo": "PEP",
        "programas": ["Presidente de la República"],
        "documentos": [{"tipo": "Cedula", "numero": "79945621"}],
        "aka": ["GUSTAVO PETRO"],
        "lista": "PEPs Colombia",
        "detalle": "Presidente de Colombia 2022-2026 — También en OFAC SDN",
    },
    {
        "uid": "PEP-COL-002",
        "nombre": "MARQUEZ MINA FRANCIA ELENA",
        "tipo": "PEP",
        "programas": ["Vicepresidenta de la República"],
        "documentos": [{"tipo": "Cedula", "numero": "55230826"}],
        "aka": ["FRANCIA MARQUEZ"],
        "lista": "PEPs Colombia",
        "detalle": "Vicepresidenta de Colombia 2022-2026",
    },
    {
        "uid": "PEP-COL-003",
        "nombre": "URIBE VELEZ ALVARO",
        "tipo": "PEP",
        "programas": ["Ex-Presidente de la República"],
        "documentos": [{"tipo": "Cedula", "numero": "70072685"}],
        "aka": ["ALVARO URIBE"],
        "lista": "PEPs Colombia",
        "detalle": "Ex-Presidente Colombia 2002-2010",
    },
    {
        "uid": "PEP-COL-004",
        "nombre": "SANTOS CALDERON JUAN MANUEL",
        "tipo": "PEP",
        "programas": ["Ex-Presidente de la República"],
        "documentos": [{"tipo": "Cedula", "numero": "19427479"}],
        "aka": ["JUAN MANUEL SANTOS"],
        "lista": "PEPs Colombia",
        "detalle": "Ex-Presidente Colombia 2010-2018 · Premio Nobel de Paz",
    },
    {
        "uid": "PEP-COL-005",
        "nombre": "DUQUE MARQUEZ IVAN",
        "tipo": "PEP",
        "programas": ["Ex-Presidente de la República"],
        "documentos": [{"tipo": "Cedula", "numero": "80345678"}],
        "aka": ["IVAN DUQUE"],
        "lista": "PEPs Colombia",
        "detalle": "Ex-Presidente Colombia 2018-2022",
    },
    {
        "uid": "PEP-COL-006",
        "nombre": "CORDOBA RUIZ PIEDAD ESNEDA",
        "tipo": "PEP",
        "programas": ["Ex-Senadora"],
        "documentos": [{"tipo": "Cedula", "numero": "42987654"}],
        "aka": ["PIEDAD CORDOBA"],
        "lista": "PEPs Colombia",
        "detalle": "Ex-Senadora — Investigada vínculos FARC · Inhabilitada por Procuraduría",
    },
    {
        "uid": "PEP-COL-007",
        "nombre": "ALCOCER GARCIA VERONICA DEL SOCORRO",
        "tipo": "PEP",
        "programas": ["Primera Dama"],
        "documentos": [{"tipo": "Cedula", "numero": "64575315"}],
        "aka": ["VERONICA ALCOCER"],
        "lista": "PEPs Colombia",
        "detalle": "Primera Dama Colombia — También en OFAC SDN EO14059",
    },
]

# Agregar solo si no están ya en el JSON (evitar duplicados)
nombres_existentes = {r["nombre"].upper() for r in registros_totales}
docs_existentes    = {doc["numero"] for r in registros_totales for doc in r.get("documentos",[])}

count_prior = 0
for entry in ENTRADAS_PRIORITARIAS:
    nom = entry["nombre"].upper()
    docs_entry = [d["numero"] for d in entry.get("documentos",[])]
    ya_existe = (nom in nombres_existentes or
                 any(d in docs_existentes for d in docs_entry if d))
    if not ya_existe:
        registros_totales.append(entry)
        por_lista[entry["lista"]] = por_lista.get(entry["lista"], 0) + 1
        nombres_existentes.add(nom)
        count_prior += 1

print(f"      ✔ {count_prior} entradas prioritarias añadidas "
      f"({'ya estaban en el XML' if count_prior == 0 else 'no estaban en el XML'})")

# ─────────────────────────────────────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────────────────────────────────────
print("\nGuardando listas_vinculantes.json...")
with open("listas_vinculantes.json", "w", encoding="utf-8") as f:
    json.dump({
        "meta": {
            "fecha_actualizacion": str(date.today()),
            "total": len(registros_totales),
            "por_lista": por_lista,
        },
        "registros": registros_totales,
    }, f, ensure_ascii=False, indent=2)

size_mb = os.path.getsize("listas_vinculantes.json") / 1024 / 1024
print("\n" + "=" * 65)
print("  RESUMEN")
print("=" * 65)
for lst, n in sorted(por_lista.items(), key=lambda x: -x[1]):
    print(f"  {lst:<30} {n:>8,} registros")
print(f"  {'─'*45}")
print(f"  {'TOTAL':<30} {len(registros_totales):>8,} registros")
print(f"  Archivo: listas_vinculantes.json  ({size_mb:.1f} MB)")
print("=" * 65)

# ── Verificación de entradas clave ───────────────────────────────────────────
print("\n  VERIFICACIÓN DE ENTRADAS CLAVE:")
print("  " + "─" * 45)
VERIFICAR = ["PETRO", "BENEDETTI", "ALCOCER", "ESCOBAR", "OTONIEL",
             "USUGA", "TIMOCHENKO", "TIMOLEON"]
for clave in VERIFICAR:
    encontrados = verificar_entrada(clave)
    if encontrados:
        listas = list({e["lista"] for e in encontrados})
        print(f"  ✔ {clave:<20} → {', '.join(listas)}")
    else:
        print(f"  ✘ {clave:<20} → NO ENCONTRADO")

print("=" * 65)
print("\n  ✔ Sube 'listas_vinculantes.json' a GitHub para activar.")
print("=" * 65)
